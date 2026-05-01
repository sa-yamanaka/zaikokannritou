# -*- coding: utf-8 -*-
"""
製造業向け 在庫管理システム
Python Flask サーバー
"""
import os
import json
import socket
import sqlite3
import io
import base64
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file, redirect
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import qrcode

# ----------------------------- 設定 -----------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "inventory.db")
HISTORY_XLSX = os.path.join(DATA_DIR, "履歴.xlsx")
CONFIG_PATH = os.path.join(BASE_DIR, "config.json")

os.makedirs(DATA_DIR, exist_ok=True)

# デフォルト設定
DEFAULT_CONFIG = {
    "port": 5000,
    "use_firebase": False,
    "firebase_config": {
        "service_account_path": "",
        "collection": "inventory"
    }
}

if not os.path.exists(CONFIG_PATH):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(DEFAULT_CONFIG, f, ensure_ascii=False, indent=2)

with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    CONFIG = json.load(f)

# ----------------------------- DB 初期化 -----------------------------
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS inventory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL UNIQUE,
            manufacturer TEXT NOT NULL,
            quantity INTEGER NOT NULL DEFAULT 0,
            length REAL,
            width REAL,
            thickness REAL,
            color TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    """)
    conn.commit()
    conn.close()

init_db()

# ----------------------------- 履歴ログ -----------------------------
def ensure_history_xlsx():
    """履歴Excelファイルが存在しなければ作成"""
    if not os.path.exists(HISTORY_XLSX):
        wb = Workbook()
        ws = wb.active
        ws.title = "増減履歴"
        headers = ["日時", "振り分け番号", "メーカー", "サイズ(縦×横×厚)", "色", "増減量", "操作後の在庫数", "操作端末"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # 列幅調整
        widths = [20, 15, 15, 20, 10, 10, 15, 12]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64+i)].width = w
        wb.save(HISTORY_XLSX)

def append_history(item, delta, new_qty, source):
    """履歴Excelに1行追加"""
    ensure_history_xlsx()
    wb = load_workbook(HISTORY_XLSX)
    ws = wb["増減履歴"]
    size = f"{item.get('length','')}×{item.get('width','')}×{item.get('thickness','')}"
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        item.get("code", ""),
        item.get("manufacturer", ""),
        size,
        item.get("color", ""),
        delta,
        new_qty,
        source
    ])
    wb.save(HISTORY_XLSX)

# ----------------------------- ルーティング -----------------------------
app = Flask(__name__, template_folder="templates", static_folder="static")

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/desktop")
def desktop():
    return render_template("desktop.html")

@app.route("/mobile")
def mobile():
    return render_template("mobile.html")

# ---------- API ----------
@app.route("/api/inventory", methods=["GET"])
def list_inventory():
    conn = get_db()
    rows = conn.execute("SELECT * FROM inventory ORDER BY code").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/inventory", methods=["POST"])
def create_inventory():
    data = request.json or {}
    code = (data.get("code") or "").strip()
    if not code:
        return jsonify({"error": "振り分け番号は必須です"}), 400
    now = datetime.now().isoformat()
    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO inventory (code, manufacturer, quantity, length, width, thickness, color, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (
            code,
            data.get("manufacturer", ""),
            int(data.get("quantity", 0) or 0),
            float(data.get("length") or 0) or None,
            float(data.get("width") or 0) or None,
            float(data.get("thickness") or 0) or None,
            data.get("color", ""),
            now, now
        ))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"error": f"振り分け番号 '{code}' は既に登録されています"}), 400
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/inventory/<int:item_id>", methods=["PUT"])
def update_inventory(item_id):
    data = request.json or {}
    now = datetime.now().isoformat()
    conn = get_db()
    conn.execute("""
        UPDATE inventory
        SET code=?, manufacturer=?, quantity=?, length=?, width=?, thickness=?, color=?, updated_at=?
        WHERE id=?
    """, (
        data.get("code"),
        data.get("manufacturer", ""),
        int(data.get("quantity", 0) or 0),
        float(data.get("length") or 0) or None,
        float(data.get("width") or 0) or None,
        float(data.get("thickness") or 0) or None,
        data.get("color", ""),
        now, item_id
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/inventory/<int:item_id>", methods=["DELETE"])
def delete_inventory(item_id):
    conn = get_db()
    conn.execute("DELETE FROM inventory WHERE id=?", (item_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/inventory/<int:item_id>/adjust", methods=["POST"])
def adjust_inventory(item_id):
    """在庫増減（履歴ログ付き）"""
    data = request.json or {}
    delta = int(data.get("delta", 0) or 0)
    source = data.get("source", "mobile")

    conn = get_db()
    row = conn.execute("SELECT * FROM inventory WHERE id=?", (item_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "対象が見つかりません"}), 404
    new_qty = (row["quantity"] or 0) + delta
    now = datetime.now().isoformat()
    conn.execute("UPDATE inventory SET quantity=?, updated_at=? WHERE id=?", (new_qty, now, item_id))
    conn.commit()
    item = dict(row)
    item["quantity"] = new_qty
    conn.close()

    # 履歴Excelに追記
    try:
        append_history(item, delta, new_qty, source)
    except Exception as e:
        print(f"[WARN] 履歴の書き込みに失敗: {e}")

    return jsonify({"ok": True, "quantity": new_qty, "alert": new_qty < 0})

@app.route("/api/sizes", methods=["GET"])
def get_sizes():
    """既存在庫に登録されているサイズ・メーカー・色の一覧を返す（プルダウン用）"""
    conn = get_db()
    rows = conn.execute("SELECT DISTINCT manufacturer FROM inventory WHERE manufacturer != '' ORDER BY manufacturer").fetchall()
    manufacturers = [r["manufacturer"] for r in rows]
    rows = conn.execute("SELECT DISTINCT length FROM inventory WHERE length IS NOT NULL ORDER BY length").fetchall()
    lengths = [r["length"] for r in rows]
    rows = conn.execute("SELECT DISTINCT width FROM inventory WHERE width IS NOT NULL ORDER BY width").fetchall()
    widths = [r["width"] for r in rows]
    rows = conn.execute("SELECT DISTINCT thickness FROM inventory WHERE thickness IS NOT NULL ORDER BY thickness").fetchall()
    thicknesses = [r["thickness"] for r in rows]
    rows = conn.execute("SELECT DISTINCT color FROM inventory WHERE color != '' ORDER BY color").fetchall()
    colors = [r["color"] for r in rows]
    conn.close()
    return jsonify({
        "manufacturers": manufacturers,
        "lengths": lengths,
        "widths": widths,
        "thicknesses": thicknesses,
        "colors": colors
    })

@app.route("/api/inventory/search", methods=["GET"])
def search_inventory():
    """検索API - code / manufacturer+size で検索"""
    code = request.args.get("code")
    manufacturer = request.args.get("manufacturer")
    length = request.args.get("length")
    width = request.args.get("width")
    thickness = request.args.get("thickness")

    conn = get_db()
    if code:
        rows = conn.execute("SELECT * FROM inventory WHERE code=?", (code,)).fetchall()
    else:
        query = "SELECT * FROM inventory WHERE 1=1"
        params = []
        if manufacturer:
            query += " AND manufacturer=?"
            params.append(manufacturer)
        if length:
            query += " AND length=?"
            params.append(float(length))
        if width:
            query += " AND width=?"
            params.append(float(width))
        if thickness:
            query += " AND thickness=?"
            params.append(float(thickness))
        rows = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/export", methods=["POST"])
def export_excel():
    """列マッピングを受け取ってExcel出力"""
    data = request.json or {}
    mapping = data.get("mapping", {})  # {"A": "code", "B": "manufacturer", ...}
    preview = data.get("preview", False)

    conn = get_db()
    rows = conn.execute("SELECT * FROM inventory ORDER BY code").fetchall()
    conn.close()

    label_map = {
        "code": "振り分け番号",
        "manufacturer": "メーカー名",
        "quantity": "在庫数",
        "length": "縦",
        "width": "横",
        "thickness": "厚み",
        "color": "色",
        "": ""
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "在庫一覧"

    # ヘッダー
    for col_letter, field_key in mapping.items():
        if field_key:
            cell = ws[f"{col_letter}1"]
            cell.value = label_map.get(field_key, field_key)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # データ
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_idx, r in enumerate(rows, start=2):
        item = dict(r)
        for col_letter, field_key in mapping.items():
            if field_key:
                cell = ws[f"{col_letter}{row_idx}"]
                val = item.get(field_key, "")
                cell.value = val
                cell.border = border
                # 在庫数マイナスを赤
                if field_key == "quantity" and isinstance(val, (int, float)) and val < 0:
                    cell.font = Font(color="FF0000", bold=True)

    # 列幅
    for col_letter, field_key in mapping.items():
        if field_key:
            ws.column_dimensions[col_letter].width = 16

    if preview:
        # プレビュー用にデータ返却
        preview_data = []
        preview_data.append([label_map.get(mapping.get(col, ""), "") for col in "ABCDEFG"])
        for r in rows[:50]:
            item = dict(r)
            row_data = []
            for col in "ABCDEFG":
                field_key = mapping.get(col, "")
                row_data.append(str(item.get(field_key, "")) if field_key else "")
            preview_data.append(row_data)
        return jsonify({"preview": preview_data, "total": len(rows)})

    # ファイルに保存して返す
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"在庫一覧_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)

@app.route("/api/history")
def download_history():
    """履歴Excelをダウンロード"""
    ensure_history_xlsx()
    return send_file(HISTORY_XLSX, as_attachment=True, download_name="増減履歴.xlsx")

@app.route("/api/qr")
def qr_code():
    """スマホ用URLのQRコードをPNGで返す"""
    ip = get_local_ip()
    url = f"http://{ip}:{CONFIG['port']}/mobile"
    img = qrcode.make(url)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")

@app.route("/api/server-info")
def server_info():
    ip = get_local_ip()
    return jsonify({
        "ip": ip,
        "port": CONFIG["port"],
        "mobile_url": f"http://{ip}:{CONFIG['port']}/mobile"
    })

# ----------------------------- ユーティリティ -----------------------------
def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

# ----------------------------- 起動 -----------------------------
if __name__ == "__main__":
    ip = get_local_ip()
    port = CONFIG["port"]
    print("=" * 60)
    print("  製造業向け 在庫管理システム")
    print("=" * 60)
    print(f"  PC用画面       : http://localhost:{port}/")
    print(f"  スマホ用URL    : http://{ip}:{port}/mobile")
    print(f"  履歴ファイル   : {HISTORY_XLSX}")
    print("=" * 60)
    print("  ブラウザを閉じてもサーバーは動き続けます。")
    print("  停止するには このウィンドウを閉じてください。")
    print("=" * 60)
    app.run(host="0.0.0.0", port=port, debug=False)
